"""
Exportação de relatórios.
  - gerar_excel()  → arquivo .xlsx com DFC do mês + evolução anual
  - gerar_pdf_ceo() → PDF executivo com KPIs, gráficos e custos de eventos
"""
from __future__ import annotations
import io
from datetime import datetime
import pandas as pd

from core.utils import fmt_br, fmt_br_kpi
from core.parser import MESES_ABREV
from core.dfc import SECOES, GRUPOS, LinhasDFC

# Paleta CPD
NAVY  = "#1C2B5F"
RED   = "#C4153A"
TEAL  = "#2A9D8F"
AMBER = "#E9A020"
GRAY  = "#F0F4F8"
WHITE = "#FFFFFF"


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def gerar_excel(
    mes: int,
    ano: int,
    dfc: LinhasDFC,
    plano_df: pd.DataFrame,
    evolucao_df: pd.DataFrame,
) -> bytes:
    """Gera .xlsx com duas abas: DFC do mês e Evolução anual."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter as gcl

    wb = Workbook()
    mes_nome = MESES_ABREV[mes]
    desc_por_cod = dict(zip(plano_df["codigo"], plano_df["descricao"]))

    def _hex(h): return h.lstrip("#")

    def _cell(ws, row, col, value="", bold=False, size=10, color=None,
               bg=None, align="left", num_fmt=None, height=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Calibri", bold=bold, size=size,
                      color=_hex(color or "#000000"))
        if bg:
            c.fill = PatternFill("solid", fgColor=_hex(bg))
        c.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=False)
        if num_fmt:
            c.number_format = num_fmt
        if height:
            ws.row_dimensions[row].height = height
        return c

    def _merge_hdr(ws, row, text, col1, col2, bg=NAVY, fg=WHITE, size=12):
        ws.merge_cells(start_row=row, start_column=col1,
                       end_row=row, end_column=col2)
        c = _cell(ws, row, col1, text, bold=True, size=size,
                  color=fg, bg=bg, align="center", height=26)
        return c

    # ── Aba 1: DFC do Mês ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = f"DFC {mes_nome}{ano}"
    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 46
    ws1.column_dimensions["C"].width = 20

    _merge_hdr(ws1, 1, f"Colégio Primeiros Degraus  |  DFC {mes_nome}/{ano}", 1, 3)
    _merge_hdr(ws1, 2, f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
               1, 3, bg=GRAY, fg=NAVY, size=9)

    r = 4
    for col, txt, al in [(1, "Código", "center"), (2, "Descrição", "left"),
                          (3, "Valor (R$)", "right")]:
        _cell(ws1, r, col, txt, bold=True, color=WHITE, bg=NAVY, align=al, height=20)
    r += 1

    for sp, si in SECOES.items():
        sec_total = dfc.total_secao(sp)
        if sp == "1.": sec_total += dfc.ar
        if sp == "3.": sec_total += dfc.ad

        _cell(ws1, r, 1, sp,            bold=True, bg="#D9E0F0", align="center", height=20)
        _cell(ws1, r, 2, si["label"],   bold=True, bg="#D9E0F0")
        _cell(ws1, r, 3, sec_total,     bold=True, bg="#D9E0F0", align="right",
              num_fmt='#,##0.00')
        r += 1

        for grp, glabel in {k: v for k, v in GRUPOS.items() if k.startswith(sp[0])}.items():
            contas = dfc.dados.get(sp, {}).get(grp, {})
            if not contas:
                continue
            grp_total = sum(contas.values())
            _cell(ws1, r, 1, grp,   bold=True, align="center", height=18)
            _cell(ws1, r, 2, "  " + glabel, bold=True)
            _cell(ws1, r, 3, grp_total, bold=True, align="right", num_fmt='#,##0.00')
            r += 1
            for cod, val in sorted(contas.items()):
                if val == 0:
                    continue
                _cell(ws1, r, 1, cod, align="center", height=16)
                _cell(ws1, r, 2, "    " + desc_por_cod.get(cod, cod))
                _cell(ws1, r, 3, val, align="right", num_fmt='#,##0.00')
                r += 1

        if sp == "1." and dfc.ar != 0:
            _cell(ws1, r, 1, "AR", align="center", height=16)
            _cell(ws1, r, 2, "  Ajuste Receita (AR)")
            _cell(ws1, r, 3, dfc.ar, align="right", num_fmt='#,##0.00')
            r += 1
        if sp == "3." and dfc.ad != 0:
            _cell(ws1, r, 1, "AD", align="center", height=16)
            _cell(ws1, r, 2, "  Ajuste Despesa (AD)")
            _cell(ws1, r, 3, dfc.ad, align="right", num_fmt='#,##0.00')
            r += 1
        r += 1

    # Resultado final
    res_bg = "#D0F0D8" if dfc.resultado_liquido >= 0 else "#F0D0D5"
    _cell(ws1, r, 1, "",                    bold=True, bg=res_bg, height=22)
    _cell(ws1, r, 2, "RESULTADO LÍQUIDO",   bold=True, bg=res_bg)
    _cell(ws1, r, 3, dfc.resultado_liquido, bold=True, bg=res_bg,
          align="right", num_fmt='#,##0.00')

    # ── Aba 2: Evolução Anual ─────────────────────────────────────────────
    if not evolucao_df.empty:
        ws2 = wb.create_sheet(title=f"Evolução {ano}")
        n_meses = len(evolucao_df)
        last_col = n_meses + 2  # 1 label + n meses + 1 total

        _merge_hdr(ws2, 1, f"Colégio Primeiros Degraus  |  Evolução Mensal {ano}",
                   1, last_col)

        # Cabeçalho colunas
        r2 = 3
        ws2.column_dimensions["A"].width = 28
        _cell(ws2, r2, 1, "Seção", bold=True, color=WHITE, bg=NAVY, height=20)
        for i, row_ev in enumerate(evolucao_df.itertuples(), 2):
            ws2.column_dimensions[gcl(i)].width = 16
            _cell(ws2, r2, i, MESES_ABREV[row_ev.mes],
                  bold=True, color=WHITE, bg=NAVY, align="right")
        ws2.column_dimensions[gcl(last_col)].width = 16
        _cell(ws2, r2, last_col, "TOTAL", bold=True, color=WHITE, bg=RED, align="right")

        r2 += 1
        fields = [
            ("receitas",  "RECEITAS"),
            ("custos",    "CUSTOS DIRETOS"),
            ("despesas",  "DESPESAS OPERACIONAIS"),
            ("impostos",  "IMPOSTOS"),
            ("resultado", "RESULTADO LÍQUIDO"),
        ]
        for field, label in fields:
            bg = "#D0F0D8" if field == "resultado" and evolucao_df["resultado"].sum() >= 0 else (
                 "#F0D0D5" if field == "resultado" else "#D9E0F0")
            vals = evolucao_df[field].tolist() if field in evolucao_df.columns else []
            _cell(ws2, r2, 1, label, bold=True, bg=bg, height=20)
            for i, v in enumerate(vals, 2):
                _cell(ws2, r2, i, float(v), bold=(field == "resultado"),
                      bg=bg, align="right", num_fmt='#,##0.00')
            _cell(ws2, r2, last_col, sum(vals), bold=True,
                  bg=bg, align="right", num_fmt='#,##0.00')
            r2 += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# PDF EXECUTIVO
# ═══════════════════════════════════════════════════════════════════════════════

def gerar_pdf_ceo(
    ano: int,
    evolucao_df: pd.DataFrame,
    dfc_ultimo: LinhasDFC,
    mes_ultimo: int,
) -> bytes:
    """Gera PDF executivo para CEOs: KPIs, gráficos e custos de eventos."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    from matplotlib.backends.backend_pdf import PdfPages

    NAVY_RGB   = (0.11, 0.17, 0.37)
    RED_RGB    = (0.77, 0.08, 0.23)
    TEAL_RGB   = (0.16, 0.61, 0.56)
    GREEN_RGB  = (0.10, 0.50, 0.22)
    LGRAY_RGB  = (0.94, 0.96, 0.97)
    WHITE_RGB  = (1.00, 1.00, 1.00)

    n_meses = len(evolucao_df)
    labels  = evolucao_df["label"].tolist() if not evolucao_df.empty else []

    def _kpi_box(ax, x, y, w, h, label, value, color, bg=None):
        """Desenha um card KPI no axes."""
        bg = bg or (*color, 0.10)
        rect = mpatches.FancyBboxPatch((x, y), w, h,
                                        boxstyle="round,pad=0.01",
                                        facecolor=bg, edgecolor=color, linewidth=1.5,
                                        transform=ax.transAxes, clip_on=False)
        ax.add_patch(rect)
        ax.text(x + 0.01, y + h - 0.01, label,
                transform=ax.transAxes, ha="left", va="top",
                fontsize=7, color=(0.3, 0.3, 0.3))
        ax.text(x + w / 2, y + h / 2 - 0.01, value,
                transform=ax.transAxes, ha="center", va="center",
                fontsize=11, fontweight="bold", color=color)

    def _page_header(fig, title: str, subtitle: str = ""):
        # Faixa azul no topo
        ax_hdr = fig.add_axes([0, 0.92, 1, 0.08])
        ax_hdr.set_facecolor(NAVY_RGB)
        ax_hdr.axis("off")
        ax_hdr.text(0.02, 0.55, "Colégio Primeiros Degraus",
                    transform=ax_hdr.transAxes, color=WHITE_RGB,
                    fontsize=14, fontweight="bold", va="center")
        ax_hdr.text(0.98, 0.55, title,
                    transform=ax_hdr.transAxes, color=(0.9, 0.9, 0.9),
                    fontsize=10, va="center", ha="right")
        if subtitle:
            ax_hdr.text(0.02, 0.15, subtitle,
                        transform=ax_hdr.transAxes, color=(0.8, 0.8, 0.9),
                        fontsize=8, va="center")

    buf = io.BytesIO()
    gerado_em = datetime.now().strftime("%d/%m/%Y %H:%M")
    periodo = f"Acumulado {ano}  —  {n_meses} {'mês' if n_meses == 1 else 'meses'} importados"

    with PdfPages(buf) as pdf:

        # ── Página 1: Resumo Executivo ────────────────────────────────────
        fig1 = plt.figure(figsize=(11, 8.5))
        fig1.patch.set_facecolor(WHITE_RGB)
        _page_header(fig1, "Relatório Executivo",
                     f"Gerado em {gerado_em}  |  {periodo}")

        # KPI cards (acumulados do ano)
        if not evolucao_df.empty:
            rec_ytd  = evolucao_df["receitas"].sum()
            sai_ytd  = evolucao_df["saidas"].sum()
            res_ytd  = evolucao_df["resultado"].sum()
            cus_ytd  = evolucao_df["custos"].sum()
            des_ytd  = evolucao_df["despesas"].sum()
            imp_ytd  = evolucao_df["impostos"].sum()
        else:
            rec_ytd = sai_ytd = res_ytd = cus_ytd = des_ytd = imp_ytd = 0.0

        res_color = GREEN_RGB if res_ytd >= 0 else RED_RGB

        ax_kpi = fig1.add_axes([0.02, 0.72, 0.96, 0.18])
        ax_kpi.axis("off")
        ax_kpi.set_xlim(0, 1); ax_kpi.set_ylim(0, 1)

        kpis = [
            ("Receitas Acumuladas",    fmt_br_kpi(rec_ytd),  TEAL_RGB),
            ("Custos Diretos",         fmt_br_kpi(cus_ytd),  RED_RGB),
            ("Despesas Operacionais",  fmt_br_kpi(des_ytd),  RED_RGB),
            ("Impostos",               fmt_br_kpi(imp_ytd),  RED_RGB),
            ("Resultado Acumulado",    fmt_br_kpi(res_ytd),  res_color),
        ]
        w_box = 0.185
        gap   = 0.0125
        for i, (lbl, val, col) in enumerate(kpis):
            x = i * (w_box + gap)
            _kpi_box(ax_kpi, x, 0.05, w_box, 0.88, lbl, val, col)

        # Tabela resumo mensal
        if not evolucao_df.empty:
            ax_tab = fig1.add_axes([0.02, 0.08, 0.96, 0.60])
            ax_tab.axis("off")

            col_labels = ["Mês", "Receitas", "Saídas", "Resultado"]
            table_data = [
                [MESES_ABREV[int(r.mes)],
                 fmt_br_kpi(r.receitas),
                 fmt_br_kpi(r.saidas),
                 fmt_br_kpi(r.resultado)]
                for r in evolucao_df.itertuples()
            ]
            # Linha de total
            table_data.append([
                "TOTAL",
                fmt_br_kpi(evolucao_df["receitas"].sum()),
                fmt_br_kpi(evolucao_df["saidas"].sum()),
                fmt_br_kpi(evolucao_df["resultado"].sum()),
            ])

            tbl = ax_tab.table(
                cellText=table_data,
                colLabels=col_labels,
                cellLoc="center",
                loc="upper center",
                colWidths=[0.15, 0.28, 0.28, 0.28],
            )
            tbl.auto_set_font_size(False)
            tbl.set_fontsize(9)
            tbl.scale(1, 1.6)

            # Header styling
            for col_i in range(4):
                cell = tbl[0, col_i]
                cell.set_facecolor(NAVY_RGB)
                cell.set_text_props(color=WHITE_RGB, fontweight="bold")

            # Total row styling
            total_row_i = len(table_data)
            for col_i in range(4):
                cell = tbl[total_row_i, col_i]
                cell.set_facecolor(LGRAY_RGB)
                cell.set_text_props(fontweight="bold")

            # Color resultado column
            for row_i in range(1, len(table_data)):
                val = evolucao_df.iloc[row_i - 1]["resultado"] if row_i <= n_meses else res_ytd
                color = (0.85, 0.96, 0.87) if val >= 0 else (0.98, 0.88, 0.88)
                tbl[row_i, 3].set_facecolor(color)

        ax_rodape = fig1.add_axes([0, 0, 1, 0.06])
        ax_rodape.axis("off")
        ax_rodape.text(0.5, 0.5, f"Colégio Primeiros Degraus  —  Relatório Confidencial  —  {gerado_em}",
                       transform=ax_rodape.transAxes, ha="center", va="center",
                       fontsize=7, color=(0.5, 0.5, 0.5))

        pdf.savefig(fig1, bbox_inches="tight")
        plt.close(fig1)

        # ── Página 2: Gráficos de Desempenho ─────────────────────────────
        if not evolucao_df.empty and n_meses >= 1:
            fig2 = plt.figure(figsize=(11, 8.5))
            fig2.patch.set_facecolor(WHITE_RGB)
            _page_header(fig2, "Desempenho Mensal", periodo)

            # Largura de barra adaptativa: máx 0.35, menor se poucos meses
            w    = min(0.35, 0.5 / max(n_meses, 3))
            x    = list(range(n_meses))
            # Margens: pelo menos 0.8 de cada lado para barras não ficarem gigantes
            margin = max(0.8, 1.5 - n_meses * 0.1)

            def _fmt_ax(v, _):
                """Formatter de eixo em R$ sem abs (mostra negativos)."""
                return f"R$ {v:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")

            # Gráfico 1: Receitas vs Saídas
            ax1 = fig2.add_axes([0.09, 0.52, 0.86, 0.36])
            bars1 = ax1.bar([i - w/2 for i in x], evolucao_df["receitas"],
                            width=w, color=[(*TEAL_RGB, 0.85)], label="Receitas")
            bars2 = ax1.bar([i + w/2 for i in x], evolucao_df["saidas"],
                            width=w, color=[(*RED_RGB, 0.85)], label="Saídas")
            ax1.set_xticks(x); ax1.set_xticklabels(labels, fontsize=9)
            ax1.set_xlim(-margin, n_meses - 1 + margin)
            ax1.set_title("Receitas × Saídas por Mês", fontsize=11, fontweight="bold",
                          color=NAVY_RGB, pad=8)
            ax1.set_facecolor(LGRAY_RGB)
            ax1.spines[["top", "right"]].set_visible(False)
            ax1.yaxis.set_major_formatter(plt.FuncFormatter(_fmt_ax))
            ax1.legend(fontsize=8)
            ax1.tick_params(axis="y", labelsize=8)

            for bar in [*bars1, *bars2]:
                h = bar.get_height()
                if h > 0:
                    ax1.text(bar.get_x() + bar.get_width()/2, h * 1.01,
                             fmt_br_kpi(h), ha="center", va="bottom", fontsize=7)

            # Gráfico 2: Resultado mensal
            ax2 = fig2.add_axes([0.09, 0.10, 0.86, 0.34])
            colors_res = [(*GREEN_RGB, 0.85) if v >= 0 else (*RED_RGB, 0.85)
                          for v in evolucao_df["resultado"]]
            bars3 = ax2.bar(x, evolucao_df["resultado"],
                            width=w * 1.8, color=colors_res)
            ax2.set_xticks(x); ax2.set_xticklabels(labels, fontsize=9)
            ax2.set_xlim(-margin, n_meses - 1 + margin)
            ax2.axhline(0, color="gray", linewidth=0.8, linestyle="--")
            ax2.set_title("Resultado Líquido por Mês", fontsize=11, fontweight="bold",
                          color=NAVY_RGB, pad=8)
            ax2.set_facecolor(LGRAY_RGB)
            ax2.spines[["top", "right"]].set_visible(False)
            ax2.yaxis.set_major_formatter(plt.FuncFormatter(_fmt_ax))
            ax2.tick_params(axis="both", labelsize=8)

            for bar in bars3:
                h = bar.get_height()
                va  = "bottom" if h >= 0 else "top"
                off = abs(h) * 0.03
                ax2.text(bar.get_x() + bar.get_width()/2,
                         h + (off if h >= 0 else -off),
                         fmt_br_kpi(h), ha="center", va=va, fontsize=8, fontweight="bold")

            ax_rodape2 = fig2.add_axes([0, 0, 1, 0.06])
            ax_rodape2.axis("off")
            ax_rodape2.text(0.5, 0.5, f"Colégio Primeiros Degraus  —  {gerado_em}",
                            transform=ax_rodape2.transAxes, ha="center",
                            fontsize=7, color=(0.5, 0.5, 0.5))

            pdf.savefig(fig2, bbox_inches="tight")
            plt.close(fig2)

        # ── Página 3: Custos de Eventos ───────────────────────────────────
        fig3 = plt.figure(figsize=(11, 8.5))
        fig3.patch.set_facecolor(WHITE_RGB)
        _page_header(fig3, "Eventos — Receitas e Custos",
                     "Comparativo de receita com eventos vs custos com eventos pedagógicos")

        # Extrai dados de eventos do DFC do último mês
        rec_eventos = sum(
            dfc_ultimo.dados.get("1.", {}).get("1.04", {}).values()
        )
        cus_eventos = abs(sum(
            dfc_ultimo.dados.get("2.", {}).get("2.04", {}).values()
        ))
        margem_eventos = rec_eventos - cus_eventos
        mes_nome_ult = MESES_ABREV[mes_ultimo]

        ax_ev_kpi = fig3.add_axes([0.02, 0.76, 0.96, 0.14])
        ax_ev_kpi.axis("off")
        ax_ev_kpi.set_xlim(0, 1); ax_ev_kpi.set_ylim(0, 1)

        ev_kpis = [
            (f"Receita com Eventos\n{mes_nome_ult}/{ano}",  fmt_br(rec_eventos),  TEAL_RGB),
            (f"Custos c/ Eventos\n{mes_nome_ult}/{ano}",    fmt_br(cus_eventos),  RED_RGB),
            (f"Margem Eventos\n{mes_nome_ult}/{ano}",       fmt_br(margem_eventos),
             GREEN_RGB if margem_eventos >= 0 else RED_RGB),
        ]
        w3 = 0.30
        g3 = 0.025
        for i, (lbl, val, col) in enumerate(ev_kpis):
            _kpi_box(ax_ev_kpi, i * (w3 + g3), 0.05, w3, 0.88, lbl, val, col)

        # Contas detalhadas de eventos (receita)
        ax_det = fig3.add_axes([0.02, 0.10, 0.96, 0.62])
        ax_det.axis("off")

        rec_contas = dfc_ultimo.dados.get("1.", {}).get("1.04", {})
        cus_contas = dfc_ultimo.dados.get("2.", {}).get("2.04", {})

        det_rows = []
        for cod, val in sorted(rec_contas.items()):
            if val != 0:
                det_rows.append([cod, "Receita", fmt_br(val)])
        for cod, val in sorted(cus_contas.items()):
            if val != 0:
                det_rows.append([cod, "Custo", fmt_br(abs(val))])

        if det_rows:
            tbl3 = ax_det.table(
                cellText=det_rows,
                colLabels=["Código", "Tipo", "Valor (R$)"],
                cellLoc="center",
                loc="upper center",
                colWidths=[0.25, 0.25, 0.30],
            )
            tbl3.auto_set_font_size(False)
            tbl3.set_fontsize(10)
            tbl3.scale(1, 2.0)
            for col_i in range(3):
                cell = tbl3[0, col_i]
                cell.set_facecolor(NAVY_RGB)
                cell.set_text_props(color=WHITE_RGB, fontweight="bold")
            for row_i in range(1, len(det_rows) + 1):
                tipo = det_rows[row_i - 1][1]
                bg = (0.85, 0.96, 0.87) if tipo == "Receita" else (0.98, 0.88, 0.88)
                for col_i in range(3):
                    tbl3[row_i, col_i].set_facecolor(bg)
        else:
            ax_det.text(0.5, 0.5, "Nenhum lançamento de evento neste mês.",
                        transform=ax_det.transAxes, ha="center", va="center",
                        fontsize=12, color=(0.5, 0.5, 0.5), style="italic")

        ax_rodape3 = fig3.add_axes([0, 0, 1, 0.06])
        ax_rodape3.axis("off")
        ax_rodape3.text(0.5, 0.5, f"Colégio Primeiros Degraus  —  {gerado_em}",
                        transform=ax_rodape3.transAxes, ha="center",
                        fontsize=7, color=(0.5, 0.5, 0.5))

        pdf.savefig(fig3, bbox_inches="tight")
        plt.close(fig3)

    return buf.getvalue()
